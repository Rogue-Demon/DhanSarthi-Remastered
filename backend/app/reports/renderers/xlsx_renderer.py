"""
XLSX Report Renderer for DhanSarthi.
Generates structured Excel workbooks with styled headers, numeric cells, auto column widths, and totals.
"""

from __future__ import annotations

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.reports.generator import ReportData


class XLSXReportRenderer:
    """Renders structured ReportData into styled Microsoft Excel (.xlsx) workbooks."""

    @staticmethod
    def render(data: ReportData) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Report"
        ws.views.sheetView[0].showGridLines = True

        # Color Palette
        brand_color = "0D9488"      # Teal
        header_fill = PatternFill(start_color=brand_color, end_color=brand_color, fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Segoe UI", size=16, bold=True, color="1E293B")
        meta_font = Font(name="Segoe UI", size=10, italic=True, color="64748B")
        section_font = Font(name="Segoe UI", size=12, bold=True, color="0F172A")
        total_font = Font(name="Segoe UI", size=11, bold=True, color="0F172A")
        body_font = Font(name="Segoe UI", size=10, color="334155")
        
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        total_top_border = Border(
            top=Side(style='thin', color='0F172A'),
            bottom=Side(style='double', color='0F172A')
        )

        row_num = 1

        # Title Block
        ws.cell(row=row_num, column=1, value="DhanSarthi Financial Statement").font = title_font
        row_num += 1
        ws.cell(row=row_num, column=1, value=data.title).font = Font(name="Segoe UI", size=13, bold=True, color="0D9488")
        row_num += 1
        ws.cell(row=row_num, column=1, value=f"Account Holder: {data.user_name} ({data.user_email})").font = meta_font
        row_num += 1
        ws.cell(row=row_num, column=1, value=f"Reporting Period: {data.period_start} to {data.period_end} | Generated: {data.generated_at}").font = meta_font
        row_num += 2

        # Summary KPIs Block
        if data.summary_kpis:
            ws.cell(row=row_num, column=1, value="Executive KPI Summary").font = section_font
            row_num += 1

            headers = ["Metric", "Value", "Context"]
            for col_idx, h in enumerate(headers, start=1):
                c = ws.cell(row=row_num, column=col_idx, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center" if col_idx > 1 else "left")
            row_num += 1

            for kpi in data.summary_kpis:
                c1 = ws.cell(row=row_num, column=1, value=kpi.label)
                c2 = ws.cell(row=row_num, column=2, value=kpi.value)
                c3 = ws.cell(row=row_num, column=3, value=kpi.subtext or "")
                for c in [c1, c2, c3]:
                    c.font = body_font
                    c.border = thin_border
                c2.font = Font(name="Segoe UI", size=10, bold=True, color="0D9488")
                c2.alignment = Alignment(horizontal="right")
                row_num += 1
            row_num += 2

        # Tables Block
        for table in data.tables:
            ws.cell(row=row_num, column=1, value=table.title).font = section_font
            row_num += 1

            # Headers
            for col_idx, h in enumerate(table.headers, start=1):
                c = ws.cell(row=row_num, column=col_idx, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="left")
            row_num += 1

            # Rows
            for row in table.rows:
                for col_idx, cell_val in enumerate(row, start=1):
                    c = ws.cell(row=row_num, column=col_idx, value=cell_val)
                    c.font = body_font
                    c.border = thin_border
                    if isinstance(cell_val, str) and (cell_val.startswith("₹") or "%" in cell_val):
                        c.alignment = Alignment(horizontal="right")
                row_num += 1

            # Totals
            if table.totals:
                for col_idx, tot_val in enumerate(table.totals, start=1):
                    c = ws.cell(row=row_num, column=col_idx, value=tot_val)
                    c.font = total_font
                    c.border = total_top_border
                    if isinstance(tot_val, str) and (tot_val.startswith("₹") or "%" in tot_val):
                        c.alignment = Alignment(horizontal="right")
                row_num += 1

            row_num += 2

        # Footnotes / Notes
        if data.footnotes:
            ws.cell(row=row_num, column=1, value="Disclaimers & Notes").font = section_font
            row_num += 1
            for note in data.footnotes:
                c = ws.cell(row=row_num, column=1, value=f"• {note}")
                c.font = meta_font
                row_num += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len and len(val_str) < 60:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
